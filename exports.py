from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from branding import logo_path_for_pdf
from config import COMPANY_AR
from database import compute_totals, get_all_settings, stage_label

BRAND_GREEN = "4CAF50"
BRAND_DARK = "263238"
HEADER_FILL = PatternFill("solid", fgColor=BRAND_GREEN)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _register_arabic_font():
    candidates = [
        Path(r"C:\Windows\Fonts\arial.ttf"),
        Path(r"C:\Windows\Fonts\tahoma.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for path in candidates:
        if path.exists():
            pdfmetrics.registerFont(TTFont("ArabicFont", str(path)))
            return "ArabicFont"
    return "Helvetica"


def _reshape(text: str) -> str:
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        return get_display(arabic_reshaper.reshape(str(text)))
    except Exception:
        return str(text)


def _company_names():
    s = get_all_settings()
    return s.get("company_name_ar", COMPANY_AR), s.get("company_name_en", "")


def _pdf_logo(max_w=4 * cm, max_h=2.5 * cm):
    path = logo_path_for_pdf()
    if not path or not path.exists() or path.suffix.lower() not in (".png", ".jpg", ".jpeg"):
        return None
    try:
        return Image(str(path), width=max_w, height=max_h)
    except Exception:
        return None


def _pdf_header_story(styles, font_name, title: str, serial: str = ""):
    name_ar, name_en = _company_names()
    story = []
    logo = _pdf_logo()
    if logo:
        story.append(logo)
        story.append(Spacer(1, 8))
    title_style = ParagraphStyle(
        "BrandTitle",
        parent=styles["Heading1"],
        fontName=font_name,
        fontSize=14,
        alignment=1,
        textColor=colors.HexColor(f"#{BRAND_GREEN}"),
    )
    sub_style = ParagraphStyle(
        "BrandSub",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        alignment=1,
        textColor=colors.HexColor("#546E7A"),
    )
    story.append(Paragraph(_reshape(name_ar), title_style))
    if name_en:
        story.append(Paragraph(name_en, sub_style))
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(_reshape(title), ParagraphStyle("DocTitle", fontName=font_name, fontSize=12, alignment=1))
    )
    if serial:
        story.append(Paragraph(_reshape(f"رقم: {serial}"), sub_style))
    story.append(Spacer(1, 14))
    return story


def _excel_brand_header(ws, title: str, serial: str = ""):
    name_ar, _ = _company_names()
    ws.sheet_view.rightToLeft = True
    ws["A1"] = name_ar
    ws["A1"].font = Font(bold=True, size=14, color=BRAND_DARK)
    ws.merge_cells("A1:E1")
    ws["A2"] = title
    ws["A2"].font = Font(bold=True, size=12)
    if serial:
        ws["A3"] = f"رقم التسلسل: {serial}"
    row_start = 5
    logo = logo_path_for_pdf()
    if logo and logo.suffix.lower() in (".png", ".jpg", ".jpeg"):
        try:
            img = XLImage(str(logo))
            img.width = 120
            img.height = 75
            ws.add_image(img, "E1")
        except Exception:
            pass
    return row_start


def export_excel(product: dict, stages: list) -> BytesIO:
    totals = compute_totals(product, stages)
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير التكلفة"
    r = _excel_brand_header(ws, f"تقرير تكلفة: {product['name']}")

    info = [
        ("المحصول", product["name"]),
        ("الوحدة", product["unit"]),
        ("الكمية", totals["quantity"]),
        ("سعر الخام / وحدة", product["raw_price_per_unit"]),
        ("إجمالي الخام", totals["raw_total"]),
    ]
    for label, val in info:
        ws.cell(r, 1, label)
        ws.cell(r, 2, val)
        r += 1

    r += 1
    for c, h in enumerate(["مرحلة", "النوع", "تكلفة/وحدة", "مبلغ ثابت", "الإجمالي"], 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r += 1
    for s in totals["stages_detail"]:
        ws.cell(r, 1, s["name"])
        ws.cell(r, 2, s["type_label"])
        ws.cell(r, 3, s["cost_per_unit"])
        ws.cell(r, 4, s["lump_sum"])
        ws.cell(r, 5, s["line_total"])
        r += 1

    r += 1
    for label, val in [
        ("إجمالي المراحل", totals["stages_total"]),
        ("إجمالي التكلفة", totals["total_cost"]),
        ("إجمالي التصدير", totals["export_total"]),
        ("الفرق", totals["margin"]),
    ]:
        ws.cell(r, 1, label)
        ws.cell(r, 2, val)
        r += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_pdf(product: dict, stages: list) -> BytesIO:
    totals = compute_totals(product, stages)
    return _generic_pdf_table(
        f"تقرير تكلفة: {product['name']}",
        [
            ["البيان", "القيمة"],
            ["الكمية", f"{totals['quantity']:,.2f}"],
            ["إجمالي الخام", f"{totals['raw_total']:,.2f}"],
            *[
                [f"{s['name']} ({s['type_label']})", f"{s['line_total']:,.2f}"]
                for s in totals["stages_detail"]
            ],
            ["إجمالي التكلفة", f"{totals['total_cost']:,.2f}"],
            ["إجمالي التصدير", f"{totals['export_total']:,.2f}"],
            ["الفرق", f"{totals['margin']:,.2f}"],
        ],
    )


def export_movements_excel(title: str, headers: list, rows: list, summary: list, serial: str = "") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير"
    r = _excel_brand_header(ws, title, serial)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r += 1
    for row in rows:
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        r += 1
    r += 1
    for label, val in summary:
        ws.cell(r, 1, label)
        ws.cell(r, 2, val)
        r += 1
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_movements_pdf(title: str, headers: list, rows: list, summary: list, serial: str = "") -> BytesIO:
    data = [headers] + rows
    for label, val in summary:
        data.append([label, val])
    return _generic_pdf_table(title, data, serial=serial)


def export_invoice_pdf(invoice: dict, items: list) -> BytesIO:
    data = [
        ["البيان", "الكمية", "السعر", "الإجمالي"],
    ]
    for it in items:
        data.append(
            [
                it["description"],
                f"{it['quantity']:,.2f}",
                f"{it['unit_price']:,.2f}",
                f"{it['line_total']:,.2f}",
            ]
        )
    data.extend(
        [
            ["", "", "المجموع", f"{invoice['subtotal']:,.2f}"],
            ["", "", "خصم", f"{invoice['discount']:,.2f}"],
            ["", "", "ضريبة", f"{invoice['tax']:,.2f}"],
            ["", "", "الإجمالي", f"{invoice['grand_total']:,.2f}"],
        ]
    )
    title = f"فاتورة {'مبيعات' if invoice['invoice_type']=='sales' else 'مشتريات'} — {invoice['party_name']}"
    return _generic_pdf_table(title, data, serial=invoice["serial_number"])


def export_invoice_excel(invoice: dict, items: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    r = _excel_brand_header(ws, f"فاتورة {invoice['serial_number']}", invoice["serial_number"])
    ws.cell(r, 1, "العميل")
    ws.cell(r, 2, invoice["party_name"])
    r += 1
    ws.cell(r, 1, "التاريخ")
    ws.cell(r, 2, invoice["invoice_date"])
    r += 2
    for c, h in enumerate(["البيان", "الكمية", "السعر", "الإجمالي"], 1):
        cell = ws.cell(r, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r += 1
    for it in items:
        ws.cell(r, 1, it["description"])
        ws.cell(r, 2, it["quantity"])
        ws.cell(r, 3, it["unit_price"])
        ws.cell(r, 4, it["line_total"])
        r += 1
    r += 1
    ws.cell(r, 1, "الإجمالي النهائي")
    ws.cell(r, 2, invoice["grand_total"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generic_pdf_table(title: str, data: list, serial: str = "") -> BytesIO:
    font_name = _register_arabic_font()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = _pdf_header_story(styles, font_name, title, serial)

    def ar_row(row):
        return [_reshape(str(c)) if isinstance(c, str) and any("\u0600" <= ch <= "\u06FF" for ch in str(c)) else str(c) for c in row]

    table_data = [ar_row(r) for r in data]
    ncol = max(len(r) for r in table_data)
    col_w = (16 * cm) / ncol
    table = Table(table_data, colWidths=[col_w] * ncol)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_GREEN}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E8F5E9")]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return buf
